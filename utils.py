import os
import PySimpleGUI as sg
from openpyxl import load_workbook

def exibir_propriedades():
    config = carregar_configuracoes()
    layout_propriedades = [
        [sg.Text('Configuração de Caminhos')],
        [sg.Text('Caminho dos Inputs:'), sg.Input(default_text=config.get('input_path', ''), key='input_path', size=(25, 1)), sg.FolderBrowse()],
        [sg.Text('Caminho dos Outputs:'), sg.Input(default_text=config.get('output_path', ''), key='output_path', size=(25, 1)), sg.FolderBrowse()],
        [sg.Button('Salvar')]
    ]

    window_propriedades = sg.Window('Propriedades', layout_propriedades)
    while True:
        event, values = window_propriedades.read()
        if event == sg.WINDOW_CLOSED:
            break
        elif event == 'Salvar':
            if os.path.exists(values['input_path']) and os.path.exists(values['output_path']):
                salvar_configuracoes(values['input_path'], values['output_path'])
                sg.popup('Configurações salvas com sucesso!')
                window_propriedades.close()
                return values['input_path'], values['output_path']
            else:
                sg.popup_error('Um ou mais diretórios não existem. Verifique os caminhos e tente novamente.')
    window_propriedades.close()
    return None, None

def salvar_configuracoes(input_path, output_path):
    with open('config.txt', 'w') as arquivo:
        arquivo.write(f"input_path={input_path}\n")
        arquivo.write(f"output_path={output_path}\n")

def carregar_configuracoes():
    config = {}
    try:
        with open('config.txt', 'r') as arquivo:
            for linha in arquivo:
                chave, valor = linha.strip().split('=', 1)
                config[chave] = valor
    except FileNotFoundError:
        pass
    return config

def load_excel_data(file_path):
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        names = [cell.value for cell in ws['F'][1:] if cell.value]
        return ws, names
    except Exception as e:
        sg.popup_error(f"Erro ao carregar dados do Excel: {e}")
        return None, []
